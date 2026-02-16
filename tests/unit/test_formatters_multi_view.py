"""Tests for multi-view rendering in output formatters."""

import csv
import io
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest
from rich.console import Console
from rich.table import Table

from graftpunk.plugins.cli_plugin import CommandResult
from graftpunk.plugins.formatters import (
    CsvFormatter,
    OutputFormatter,
    TableFormatter,
    XlsxFormatter,
    format_output,
)


def _parse_csv_output(console: MagicMock) -> list[list[str]]:
    """Extract and parse CSV output from a mock console."""
    output = console.print.call_args[0][0]
    return list(csv.reader(io.StringIO(output)))


class TestTableFormatterMultiView:
    """Tests for TableFormatter multi-view rendering."""

    def test_single_view_no_section_header(self) -> None:
        """A single view renders without a Rule separator."""
        from rich.rule import Rule

        from graftpunk.plugins import OutputConfig, ViewConfig

        console = MagicMock(spec=Console)
        data = {"results": [{"id": 1, "name": "foo"}]}
        cfg = OutputConfig(
            views=[ViewConfig(name="items", path="results", title="Items")],
        )
        TableFormatter().format(data, console, output_config=cfg)
        # Should print exactly one thing: the table
        console.print.assert_called_once()
        table = console.print.call_args[0][0]
        assert isinstance(table, Table)
        # No Rule should have been printed
        for call in console.print.call_args_list:
            args = call[0]
            for arg in args:
                assert not isinstance(arg, Rule)

    def test_multiple_views_renders_all_with_headers(self) -> None:
        """Multiple views each get a Rule header and a Table."""
        from rich.rule import Rule

        from graftpunk.plugins import OutputConfig, ViewConfig

        console = MagicMock(spec=Console)
        data = {
            "items": [{"id": 1, "name": "foo"}],
            "page": {"current": 1, "total": 5},
        }
        cfg = OutputConfig(
            views=[
                ViewConfig(name="items", path="items", title="Results"),
                ViewConfig(name="page", path="page", title="Pagination"),
            ],
        )
        TableFormatter().format(data, console, output_config=cfg)
        # Collect all printed args
        rules = []
        tables = []
        for call in console.print.call_args_list:
            args = call[0]
            for arg in args:
                if isinstance(arg, Rule):
                    rules.append(arg)
                elif isinstance(arg, Table):
                    tables.append(arg)
        assert len(rules) == 2
        assert len(tables) == 2
        assert rules[0].title == "Results"
        assert rules[1].title == "Pagination"

    def test_empty_view_data_skipped(self) -> None:
        """A view whose path yields None is silently skipped."""
        from graftpunk.plugins import OutputConfig, ViewConfig

        console = MagicMock(spec=Console)
        data = {"items": [{"id": 1}]}
        cfg = OutputConfig(
            views=[
                ViewConfig(name="items", path="items"),
                ViewConfig(name="missing", path="nonexistent"),
            ],
        )
        TableFormatter().format(data, console, output_config=cfg)
        # Only one table rendered (the "missing" view is skipped)
        tables = [
            call[0][0]
            for call in console.print.call_args_list
            if call[0] and isinstance(call[0][0], Table)
        ]
        assert len(tables) == 1

    def test_multi_view_with_column_filter(self) -> None:
        """Column filters apply per-view in multi-view mode."""
        from graftpunk.plugins import ColumnFilter, OutputConfig, ViewConfig

        console = MagicMock(spec=Console)
        data = {
            "items": [{"id": 1, "name": "foo", "desc": "long"}],
        }
        cfg = OutputConfig(
            views=[
                ViewConfig(
                    name="items",
                    path="items",
                    columns=ColumnFilter("include", ["id", "name"]),
                ),
            ],
        )
        TableFormatter().format(data, console, output_config=cfg)
        console.print.assert_called_once()
        table = console.print.call_args[0][0]
        assert isinstance(table, Table)
        assert len(table.columns) == 2

    def test_no_config_still_works(self) -> None:
        """No output_config renders same as before (backward compat)."""
        console = MagicMock(spec=Console)
        data = [{"id": 1, "name": "foo"}]
        TableFormatter().format(data, console, output_config=None)
        console.print.assert_called_once()
        table = console.print.call_args[0][0]
        assert isinstance(table, Table)
        assert len(table.columns) == 2
        assert table.row_count == 1

    def test_single_dict_view_renders_key_value_pairs(self) -> None:
        """A view that extracts a single dict renders as key-value table."""
        from graftpunk.plugins import OutputConfig, ViewConfig

        console = MagicMock(spec=Console)
        data = {"meta": {"version": "1.0", "status": "ok"}}
        cfg = OutputConfig(
            views=[ViewConfig(name="meta", path="meta")],
        )
        TableFormatter().format(data, console, output_config=cfg)
        console.print.assert_called_once()
        table = console.print.call_args[0][0]
        assert isinstance(table, Table)
        assert table.row_count == 2

    def test_title_fallback_to_name(self) -> None:
        """Multi-view Rule header falls back to view name when title is empty."""
        from rich.rule import Rule

        from graftpunk.plugins import OutputConfig, ViewConfig

        console = MagicMock(spec=Console)
        data = {
            "items": [{"id": 1}],
            "page": [{"number": 0}],
        }
        cfg = OutputConfig(
            views=[
                ViewConfig(name="items", path="items"),
                ViewConfig(name="page", path="page"),
            ],
        )
        TableFormatter().format(data, console, output_config=cfg)
        rules = [
            c[0][0] for c in console.print.call_args_list if c[0] and isinstance(c[0][0], Rule)
        ]
        assert len(rules) == 2
        assert rules[0].title == "items"
        assert rules[1].title == "page"

    def test_all_views_empty_logs_warning(self) -> None:
        """When all views yield None data, a warning is logged."""
        from graftpunk.plugins import OutputConfig, ViewConfig

        console = MagicMock(spec=Console)
        data = {"other": "value"}
        cfg = OutputConfig(
            views=[
                ViewConfig(name="a", path="missing_a"),
                ViewConfig(name="b", path="missing_b"),
            ],
        )
        with patch("graftpunk.plugins.formatters.LOG") as mock_log:
            TableFormatter().format(data, console, output_config=cfg)
            mock_log.warning.assert_called_once()
            assert mock_log.warning.call_args[0][0] == "multi_view_all_empty"


class TestCsvFormatterMultiView:
    """Tests for CsvFormatter behavior with multi-view OutputConfig."""

    def test_warns_when_multiple_views_unfiltered(self) -> None:
        """CsvFormatter warns to stderr when OutputConfig has >1 view."""
        from graftpunk.plugins import OutputConfig, ViewConfig

        console = MagicMock(spec=Console)
        data = {
            "items": [{"id": 1, "name": "foo"}],
            "page": [{"number": 0}],
        }
        cfg = OutputConfig(
            views=[
                ViewConfig(name="items", path="items", title="Items"),
                ViewConfig(name="page", path="page", title="Page Info"),
            ],
            default_view="items",
        )
        with patch("graftpunk.plugins.formatters.gp_console") as mock_gp:
            CsvFormatter().format(data, console, output_config=cfg)
            mock_gp.warn.assert_called_once()
            warning_msg = mock_gp.warn.call_args[0][0]
            assert "items" in warning_msg
            assert "page" in warning_msg
            assert "--view" in warning_msg

    def test_no_warning_for_single_view(self) -> None:
        """CsvFormatter does not warn when OutputConfig has only 1 view."""
        from graftpunk.plugins import OutputConfig, ViewConfig

        console = MagicMock(spec=Console)
        data = [{"id": 1, "name": "foo"}]
        cfg = OutputConfig(views=[ViewConfig(name="items", title="Items")])
        with patch("graftpunk.plugins.formatters.gp_console") as mock_gp:
            CsvFormatter().format(data, console, output_config=cfg)
            mock_gp.warn.assert_not_called()

    def test_renders_default_view_when_multiple_views(self) -> None:
        """CsvFormatter renders the default view data, not all views."""
        from graftpunk.plugins import OutputConfig, ViewConfig

        console = MagicMock(spec=Console)
        data = {
            "items": [{"id": 1, "name": "foo"}],
            "page": [{"number": 0}],
        }
        cfg = OutputConfig(
            views=[
                ViewConfig(name="items", path="items", title="Items"),
                ViewConfig(name="page", path="page", title="Page Info"),
            ],
            default_view="items",
        )
        with patch("graftpunk.plugins.formatters.gp_console"):
            CsvFormatter().format(data, console, output_config=cfg)
        rows = _parse_csv_output(console)
        assert rows[0] == ["id", "name"]
        assert rows[1] == ["1", "foo"]


class TestXlsxFormatter:
    """Tests for the XlsxFormatter."""

    def test_satisfies_protocol(self) -> None:
        """XlsxFormatter satisfies the OutputFormatter protocol."""
        assert isinstance(XlsxFormatter(), OutputFormatter)

    def test_writes_xlsx_file(self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
        """XlsxFormatter creates a .xlsx file in the downloads directory."""
        console = MagicMock(spec=Console)
        data = [{"id": 1, "name": "foo"}, {"id": 2, "name": "bar"}]
        monkeypatch.setenv("GP_DOWNLOADS_DIR", str(tmp_path))
        with patch("graftpunk.plugins.formatters.gp_console") as mock_gp:
            XlsxFormatter().format(data, console)
            mock_gp.info.assert_called_once()
            output_msg = mock_gp.info.call_args[0][0]
            assert str(tmp_path) in output_msg
            assert ".xlsx" in output_msg
        xlsx_files = list(tmp_path.glob("*.xlsx"))
        assert len(xlsx_files) == 1

    def test_single_view_creates_named_worksheet(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """A single view creates one worksheet with the view title."""
        from graftpunk.plugins import OutputConfig, ViewConfig

        console = MagicMock(spec=Console)
        data = {"items": [{"id": 1, "name": "foo"}]}
        cfg = OutputConfig(
            views=[ViewConfig(name="items", path="items", title="Products")],
        )
        monkeypatch.setenv("GP_DOWNLOADS_DIR", str(tmp_path))
        with patch("graftpunk.plugins.formatters.gp_console"):
            XlsxFormatter().format(data, console, output_config=cfg)
        xlsx_files = list(tmp_path.glob("*.xlsx"))
        assert len(xlsx_files) == 1

    def test_multiple_views_creates_multiple_worksheets(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Multiple views each become a separate worksheet."""
        from graftpunk.plugins import OutputConfig, ViewConfig

        console = MagicMock(spec=Console)
        data = {
            "items": [{"id": 1, "name": "foo"}],
            "page": [{"number": 0, "totalPages": 5}],
        }
        cfg = OutputConfig(
            views=[
                ViewConfig(name="items", path="items", title="Products"),
                ViewConfig(name="page", path="page", title="Page Info"),
            ],
        )
        monkeypatch.setenv("GP_DOWNLOADS_DIR", str(tmp_path))
        with patch("graftpunk.plugins.formatters.gp_console"):
            XlsxFormatter().format(data, console, output_config=cfg)
        xlsx_files = list(tmp_path.glob("*.xlsx"))
        assert len(xlsx_files) == 1

    def test_empty_data_still_creates_file(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """An empty list still produces a valid .xlsx file."""
        console = MagicMock(spec=Console)
        monkeypatch.setenv("GP_DOWNLOADS_DIR", str(tmp_path))
        with patch("graftpunk.plugins.formatters.gp_console") as mock_gp:
            XlsxFormatter().format([], console)
            mock_gp.info.assert_called_once()
            output_msg = mock_gp.info.call_args[0][0]
            assert ".xlsx" in output_msg

    def test_single_dict_renders_as_key_value(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """A plain dict is rendered as key-value pairs."""
        console = MagicMock(spec=Console)
        data = {"name": "Alice", "age": 30}
        monkeypatch.setenv("GP_DOWNLOADS_DIR", str(tmp_path))
        with patch("graftpunk.plugins.formatters.gp_console"):
            XlsxFormatter().format(data, console)
        xlsx_files = list(tmp_path.glob("*.xlsx"))
        assert len(xlsx_files) == 1

    def test_xlsx_in_builtin_formatters(self) -> None:
        """XlsxFormatter is registered in BUILTIN_FORMATTERS."""
        from graftpunk.plugins.formatters import BUILTIN_FORMATTERS

        assert "xlsx" in BUILTIN_FORMATTERS

    def test_column_filter_applied(self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
        """Column filters from OutputConfig are applied to xlsx output."""
        from graftpunk.plugins import ColumnFilter, OutputConfig, ViewConfig

        console = MagicMock(spec=Console)
        data = [{"id": 1, "name": "foo", "desc": "long"}]
        cfg = OutputConfig(
            views=[
                ViewConfig(
                    name="default",
                    columns=ColumnFilter("include", ["id", "name"]),
                ),
            ],
        )
        monkeypatch.setenv("GP_DOWNLOADS_DIR", str(tmp_path))
        with patch("graftpunk.plugins.formatters.gp_console"):
            XlsxFormatter().format(data, console, output_config=cfg)
        xlsx_files = list(tmp_path.glob("*.xlsx"))
        assert len(xlsx_files) == 1

    def test_list_of_dicts_content(self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
        """Verify XLSX cell content for a list of dicts."""
        import openpyxl

        console = MagicMock(spec=Console)
        data = [{"id": 1, "name": "Alice"}, {"id": 2, "name": "Bob"}]
        monkeypatch.setenv("GP_DOWNLOADS_DIR", str(tmp_path))
        with patch("graftpunk.plugins.formatters.gp_console"):
            XlsxFormatter().format(data, console)
        xlsx_file = next(tmp_path.glob("*.xlsx"))
        wb = openpyxl.load_workbook(xlsx_file)
        ws = wb.active
        assert ws.cell(1, 1).value == "id"
        assert ws.cell(1, 2).value == "name"
        assert ws.cell(2, 1).value == 1
        assert ws.cell(2, 2).value == "Alice"
        assert ws.cell(3, 1).value == 2
        assert ws.cell(3, 2).value == "Bob"
        wb.close()

    def test_multi_view_worksheet_names(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Multiple views create worksheets with correct names and content."""
        import openpyxl

        from graftpunk.plugins import OutputConfig, ViewConfig

        console = MagicMock(spec=Console)
        data = {
            "items": [{"id": 1, "name": "foo"}],
            "meta": {"version": "2.0"},
        }
        cfg = OutputConfig(
            views=[
                ViewConfig(name="items", path="items", title="Products"),
                ViewConfig(name="meta", path="meta", title="Metadata"),
            ],
        )
        monkeypatch.setenv("GP_DOWNLOADS_DIR", str(tmp_path))
        with patch("graftpunk.plugins.formatters.gp_console"):
            XlsxFormatter().format(data, console, output_config=cfg)
        xlsx_file = next(tmp_path.glob("*.xlsx"))
        wb = openpyxl.load_workbook(xlsx_file)
        assert wb.sheetnames == ["Products", "Metadata"]
        ws_items = wb["Products"]
        assert ws_items.cell(1, 1).value == "id"
        assert ws_items.cell(2, 1).value == 1
        ws_meta = wb["Metadata"]
        assert ws_meta.cell(1, 1).value == "Key"
        assert ws_meta.cell(2, 1).value == "version"
        assert ws_meta.cell(2, 2).value == "2.0"
        wb.close()

    def test_sheet_name_truncated_to_31_chars(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """Worksheet name is truncated to Excel's 31-character limit."""
        import openpyxl

        from graftpunk.plugins import OutputConfig, ViewConfig

        console = MagicMock(spec=Console)
        long_title = "A" * 50
        data = {"items": [{"id": 1}]}
        cfg = OutputConfig(
            views=[ViewConfig(name="items", path="items", title=long_title)],
        )
        monkeypatch.setenv("GP_DOWNLOADS_DIR", str(tmp_path))
        with patch("graftpunk.plugins.formatters.gp_console"):
            XlsxFormatter().format(data, console, output_config=cfg)
        xlsx_file = next(tmp_path.glob("*.xlsx"))
        wb = openpyxl.load_workbook(xlsx_file)
        assert wb.sheetnames == ["A" * 31]
        wb.close()

    def test_empty_view_skipped(self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
        """Views whose path yields None are skipped (no empty worksheet)."""
        import openpyxl

        from graftpunk.plugins import OutputConfig, ViewConfig

        console = MagicMock(spec=Console)
        data = {"items": [{"id": 1}]}
        cfg = OutputConfig(
            views=[
                ViewConfig(name="items", path="items", title="Items"),
                ViewConfig(name="missing", path="nonexistent", title="Missing"),
            ],
        )
        monkeypatch.setenv("GP_DOWNLOADS_DIR", str(tmp_path))
        with patch("graftpunk.plugins.formatters.gp_console"):
            XlsxFormatter().format(data, console, output_config=cfg)
        xlsx_file = next(tmp_path.glob("*.xlsx"))
        wb = openpyxl.load_workbook(xlsx_file)
        assert wb.sheetnames == ["Items"]
        wb.close()


class TestFormatOutputViewArgs:
    """Tests for format_output with --view filtering."""

    def test_view_args_filters_to_single_view(self) -> None:
        """format_output with view_args filters OutputConfig to one view."""
        from graftpunk.plugins import OutputConfig, ViewConfig

        console = MagicMock(spec=Console)
        data = {
            "items": [{"id": 1, "name": "foo"}],
            "page": [{"number": 0}],
        }
        cfg = OutputConfig(
            views=[
                ViewConfig(name="items", path="items", title="Items"),
                ViewConfig(name="page", path="page", title="Page Info"),
            ],
        )
        result = CommandResult(data=data, output_config=cfg, format_hint="table")
        format_output(result, "table", console, view_args=("items",))
        # Should render only one table (items), not two
        printed_args = [c[0][0] for c in console.print.call_args_list]
        tables = [a for a in printed_args if isinstance(a, Table)]
        assert len(tables) == 1

    def test_view_args_with_column_override(self) -> None:
        """format_output with view_args applies column override."""
        from graftpunk.plugins import OutputConfig, ViewConfig

        console = MagicMock(spec=Console)
        data = [{"id": 1, "name": "foo", "desc": "long"}]
        cfg = OutputConfig(
            views=[ViewConfig(name="default")],
        )
        result = CommandResult(data=data, output_config=cfg, format_hint="table")
        format_output(result, "table", console, view_args=("default:id,name",))
        printed_args = [c[0][0] for c in console.print.call_args_list]
        tables = [a for a in printed_args if isinstance(a, Table)]
        assert len(tables) == 1
        assert len(tables[0].columns) == 2  # Only id, name

    def test_view_args_empty_tuple_renders_all(self) -> None:
        """Empty view_args renders all views (default behavior)."""
        from graftpunk.plugins import OutputConfig, ViewConfig

        console = MagicMock(spec=Console)
        data = {
            "items": [{"id": 1}],
            "page": [{"number": 0}],
        }
        cfg = OutputConfig(
            views=[
                ViewConfig(name="items", path="items", title="Items"),
                ViewConfig(name="page", path="page", title="Page Info"),
            ],
        )
        result = CommandResult(data=data, output_config=cfg, format_hint="table")
        format_output(result, "table", console, view_args=())
        printed_args = [c[0][0] for c in console.print.call_args_list]
        tables = [a for a in printed_args if isinstance(a, Table)]
        assert len(tables) == 2

    def test_view_args_no_output_config_warns(self) -> None:
        """view_args with no output_config warns and still renders."""
        console = MagicMock(spec=Console)
        result = CommandResult(data={"key": "value"})
        with patch("graftpunk.plugins.formatters.gp_console") as mock_gp:
            format_output(result, "table", console, view_args=("items",))
            mock_gp.warn.assert_called_once()
        console.print.assert_called()

    def test_nonexistent_view_name_logs_warning(self) -> None:
        """format_output with a nonexistent view name triggers a warning log."""
        from graftpunk.plugins import OutputConfig, ViewConfig

        console = MagicMock(spec=Console)
        cfg = OutputConfig(views=[ViewConfig(name="items", path="items")])
        result = CommandResult(
            data={"items": [{"id": 1}]},
            output_config=cfg,
            format_hint="table",
        )
        with patch("graftpunk.plugins.output_config.LOG") as mock_log:
            format_output(result, "table", console, view_args=("nonexistent",))
            mock_log.warning.assert_called()
            assert mock_log.warning.call_args[0][0] == "filter_views_unknown"

    def test_backward_compat_no_view_args(self) -> None:
        """format_output without view_args still works (backward compat)."""
        console = MagicMock(spec=Console)
        result = CommandResult(data={"key": "value"})
        format_output(result, "json", console)
        console.print.assert_called_once()
